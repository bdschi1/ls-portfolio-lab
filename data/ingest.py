"""Portfolio import from Excel, CSV, and PDF files.

Reads user-provided portfolio files and converts them into a Portfolio model.
Handles various column naming conventions and fills in missing data.
"""

from __future__ import annotations

import logging
import re
from collections.abc import Callable
from datetime import date
from pathlib import Path

import polars as pl

from core.portfolio import Portfolio, Position
from data.sector_map import classify_ticker

logger = logging.getLogger(__name__)

PriceFetcher = Callable[[list[str]], dict[str, float]]

# Common column name aliases — map to our standard names
COLUMN_ALIASES: dict[str, list[str]] = {
    "ticker": ["ticker", "symbol", "sym", "stock", "name", "security", "equity"],
    "side": ["side", "direction", "long_short", "l/s", "ls", "position_type"],
    "shares": ["shares", "quantity", "qty", "size", "units", "position"],
    "notional_usd": [
        "notional_usd",
        "dollar_amount",
        "amount",
        "notional",
        "dollars",
        "usd",
        "mv",
        "market_value",
        "position_size",
        "position_value",
        "exposure",
        "net_value",
        "gross_value",
    ],
    "entry_price": ["entry_price", "entry", "avg_price", "avg_cost", "cost", "price", "avg"],
    "entry_date": ["entry_date", "date", "trade_date", "open_date", "start_date"],
    "sector": ["sector", "sector_override", "gics_sector"],
    "subsector": ["subsector", "industry", "sub_sector", "subsector_override"],
    "weight": ["weight", "wt", "pct", "allocation", "weight_pct"],
    "asset_type": ["asset_type", "type", "instrument", "instrument_type", "product_type"],
    "rsi": ["rsi", "relative_strength", "relative_strength_index"],
    "beta": ["beta", "beta_to_spy", "spy_beta", "bbg_beta"],
}

# Valid side labels → standardized
SIDE_ALIASES: dict[str, str] = {
    "long": "LONG",
    "l": "LONG",
    "buy": "LONG",
    "1": "LONG",
    "+1": "LONG",
    "short": "SHORT",
    "s": "SHORT",
    "sell": "SHORT",
    "-1": "SHORT",
}


# Bloomberg terminal ticker patterns:
#   - Full:  'AAPL US EQUITY' (2-letter exchange code)
#   - Bare:  'AAPL EQUITY'    (single ticker token + EQUITY, no exchange)
# 'AAPL U EQUITY' / 'AAPL USA EQUITY' (1- or 3-letter middle token) are NOT touched.
_BLOOMBERG_FULL_SUFFIX_RE = re.compile(r"\s+[A-Z]{2}\s+EQUITY$")
_BLOOMBERG_BARE_SUFFIX_RE = re.compile(r"^([A-Z0-9.\-]+)\s+EQUITY$")
_HEADER_NORMALIZE_RE = re.compile(r"[^a-z0-9]+")

# Unit markers in headers: '($M)', '(MM)', '$K', '(millions)', '(000s)', etc.
_UNIT_MARKER_RE = re.compile(
    r"\(\s*\$?\s*(mm|millions?|m|bn|billions?|b|thousands?|000s?|k)\s*\)\s*$|"
    r"\s*\$\s*(mm|millions?|m|bn|billions?|b|thousands?|000s?|k)\s*$",
    re.IGNORECASE,
)
_UNIT_SCALES: dict[str, float] = {
    "k": 1e3,
    "thousand": 1e3,
    "thousands": 1e3,
    "000": 1e3,
    "000s": 1e3,
    "m": 1e6,
    "mm": 1e6,
    "million": 1e6,
    "millions": 1e6,
    "b": 1e9,
    "bn": 1e9,
    "billion": 1e9,
    "billions": 1e9,
}
# Tabs whose names match these are never treated as the portfolio tab.
_TAB_NAME_BLOCKLIST = re.compile(
    r"^\s*(summary|dashboard|notes?|legend|key|cover|toc|index|charts?|graphs?|"
    r"history|history\s*log|trade\s*log|metadata|settings?)\b",
    re.IGNORECASE,
)
_MAX_HEADER_SCAN_ROWS = 15


def _clean_ticker(raw: str) -> str:
    """Strip Bloomberg terminal suffixes (e.g., 'AAPL US EQUITY' -> 'AAPL').

    Handles 'TICKER <exchange> EQUITY' (2-letter exchange: US, LN, JP, GR, FP, CN, AU, ...)
    and bare 'TICKER EQUITY' (no exchange). Tickers with a non-2-letter middle token
    (e.g. 'AAPL USA EQUITY') are left unchanged. Plain tickers pass through unchanged.
    """
    cleaned = raw.strip().upper()
    cleaned = _BLOOMBERG_FULL_SUFFIX_RE.sub("", cleaned)
    bare_match = _BLOOMBERG_BARE_SUFFIX_RE.match(cleaned)
    if bare_match:
        cleaned = bare_match.group(1)
    return cleaned


def _extract_unit_scale(header: str) -> tuple[str, float]:
    """Strip trailing unit markers like '($M)', '($MM)', '$K' from a header.

    Returns (header_without_unit, scale_factor). Scale factor is 1.0 if no marker.
    """
    if not isinstance(header, str):
        return str(header) if header is not None else "", 1.0
    match = _UNIT_MARKER_RE.search(header)
    if not match:
        return header, 1.0
    unit = (match.group(1) or match.group(2) or "").lower()
    scale = _UNIT_SCALES.get(unit, 1.0)
    cleaned = header[: match.start()].rstrip()
    return cleaned, scale


def _normalize_header(name: str) -> str:
    """Lowercase, drop $, collapse non-alphanumerics to '_', trim leading/trailing '_'."""
    no_dollar = name.replace("$", "")
    return _HEADER_NORMALIZE_RE.sub("_", no_dollar.lower()).strip("_")


def _standardize_columns(df: pl.DataFrame) -> tuple[pl.DataFrame, dict[str, float]]:
    """Rename columns to standard names using alias lookup.

    Detects unit markers ('($M)', '$K', etc.) on column headers, strips them,
    and returns a {standard_column_name: scale_factor} dict for downstream scaling.
    Columns whose original header contains a literal '$' (e.g. '$ Amount',
    'Position Size ($M)') are force-routed to 'notional_usd' even when the
    stem alias would otherwise match 'shares'.
    """
    rename_map: dict[str, str] = {}
    scales: dict[str, float] = {}

    # Build {normalized_clean_header: (original_header, scale, had_dollar)}
    parsed: dict[str, tuple[str, float, bool]] = {}
    for col in df.columns:
        cleaned, scale = _extract_unit_scale(col)
        normalized = _normalize_header(cleaned)
        had_dollar = "$" in col
        if normalized:
            parsed[normalized] = (col, scale, had_dollar)

    # First pass: force-route any $-marked column to notional_usd
    for normalized, (original, scale, had_dollar) in parsed.items():
        if had_dollar and "notional_usd" not in rename_map.values():
            rename_map[original] = "notional_usd"
            scales["notional_usd"] = scale

    # Second pass: normal alias matching for everything else
    for standard, aliases in COLUMN_ALIASES.items():
        if standard in rename_map.values():
            continue
        for alias in aliases:
            if alias in parsed:
                original, scale, _ = parsed[alias]
                if original in rename_map:
                    continue
                rename_map[original] = standard
                if scale != 1.0:
                    scales[standard] = scale
                break

    return df.rename(rename_map), scales


def _standardize_side(side_val: str) -> str:
    """Convert various side labels to LONG or SHORT."""
    cleaned = str(side_val).strip().lower()
    result = SIDE_ALIASES.get(cleaned)
    if result is None:
        msg = f"Cannot interpret side value: '{side_val}'. Expected LONG/SHORT/L/S/BUY/SELL."
        raise ValueError(msg)
    return result


def _row_matches_portfolio_header(headers: list[object]) -> bool:
    """A row qualifies as a portfolio header if it contains both a ticker-alias
    column and any sizing column (shares / notional / weight, or any $-marked column)."""
    string_headers = [h for h in headers if isinstance(h, str) and h.strip()]
    if not string_headers:
        return False
    normalized = {_normalize_header(_extract_unit_scale(h)[0]) for h in string_headers}
    has_ticker = any(alias in normalized for alias in COLUMN_ALIASES["ticker"])
    has_size = (
        any(alias in normalized for alias in COLUMN_ALIASES["shares"])
        or any(alias in normalized for alias in COLUMN_ALIASES["notional_usd"])
        or any(alias in normalized for alias in COLUMN_ALIASES["weight"])
        or any("$" in h for h in string_headers)
    )
    return has_ticker and has_size


def _read_sheet_to_df(workbook, sheet_name: str, header_row: int) -> pl.DataFrame:
    """Read a worksheet into a Polars DataFrame using a specific 1-based header row."""
    ws = workbook[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    if header_row > len(all_rows):
        msg = f"Sheet {sheet_name!r} has fewer than {header_row} rows."
        raise ValueError(msg)

    raw_headers = all_rows[header_row - 1]
    data_rows = all_rows[header_row:]

    # Build columns, fabricating placeholder names for blank header cells
    headers: list[str] = []
    keep_cols: list[int] = []
    for i, h in enumerate(raw_headers):
        if h is None or (isinstance(h, str) and not h.strip()):
            continue
        headers.append(str(h).strip())
        keep_cols.append(i)

    columns: dict[str, list] = {h: [] for h in headers}
    for row in data_rows:
        # Drop trailing all-None rows
        if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
            continue
        for h, col_idx in zip(headers, keep_cols, strict=True):
            columns[h].append(row[col_idx] if col_idx < len(row) else None)

    # All columns must be equal length; trim/pad just in case
    if columns:
        n = max(len(v) for v in columns.values())
        for h in columns:
            if len(columns[h]) < n:
                columns[h] += [None] * (n - len(columns[h]))

    return pl.DataFrame(columns, strict=False)


def _find_portfolio_sheet_and_header(workbook) -> tuple[str, int] | None:
    """Scan every sheet for a row that looks like a portfolio header.

    Returns (sheet_name, header_row_1based) or None if nothing matches.
    Sheets whose names match the dashboard/summary blocklist are tried last
    so a real 'Portfolio' tab wins over a 'Summary' tab.
    """
    candidates = list(workbook.sheetnames)
    deprioritized = [s for s in candidates if _TAB_NAME_BLOCKLIST.match(s)]
    preferred = [s for s in candidates if s not in deprioritized]

    for sheet_name in preferred + deprioritized:
        ws = workbook[sheet_name]
        rows_to_scan = min(_MAX_HEADER_SCAN_ROWS, ws.max_row or 0)
        for header_row in range(1, rows_to_scan + 1):
            row_values = [cell.value for cell in ws[header_row]]
            if _row_matches_portfolio_header(row_values):
                return sheet_name, header_row

    return None


def load_from_excel(
    file_path: str | Path,
    nav: float = 3_000_000_000.0,
    sheet_name: str | None = None,
    price_fetcher: PriceFetcher | None = None,
) -> Portfolio:
    """
    Load a portfolio from an Excel (.xlsx) or CSV (.csv) file.

    Auto-discovers the portfolio tab and header row when the workbook has multiple
    sheets, summary headers, or banner rows above the table. Pass `sheet_name`
    explicitly to skip the discovery scan.

    Required columns: ticker + one of (shares / weight / a $ notional column).
    Side column is optional when shares/notional are signed.
    Headers may carry unit markers like '($M)', '$MM', '$K' — values are scaled
    accordingly. Headers with a literal '$' (e.g. 'Position Size ($M)') are
    routed to the dollar-notional path.

    When notional is supplied without entry_price and `price_fetcher` is provided,
    shares are derived as notional / current_price.

    Bloomberg-format tickers ('AAPL US EQUITY', 'AAPL EQUITY') are auto-cleaned.
    """
    path = Path(file_path)

    if path.suffix.lower() == ".csv":
        df = pl.read_csv(path, infer_schema_length=1000)
        return _parse_portfolio_df(df, nav=nav, source=str(path), price_fetcher=price_fetcher)

    if path.suffix.lower() not in (".xlsx", ".xls"):
        msg = f"Unsupported file type: {path.suffix}. Use .csv, .xlsx, or .xls."
        raise ValueError(msg)

    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name is not None:
            # Caller specified the sheet; find header row within it (default to row 1)
            ws = workbook[sheet_name]
            header_row = 1
            for candidate in range(1, min(_MAX_HEADER_SCAN_ROWS, ws.max_row or 0) + 1):
                row_values = [cell.value for cell in ws[candidate]]
                if _row_matches_portfolio_header(row_values):
                    header_row = candidate
                    break
            df = _read_sheet_to_df(workbook, sheet_name, header_row)
            return _parse_portfolio_df(
                df,
                nav=nav,
                source=f"{path}::{sheet_name}",
                price_fetcher=price_fetcher,
            )

        discovered = _find_portfolio_sheet_and_header(workbook)
        if discovered is None:
            msg = (
                f"Could not find a portfolio table in {path.name}. "
                f"Looked at sheets: {workbook.sheetnames}. "
                f"Expected a row with a ticker column plus a sizing column "
                f"(shares / notional / weight / $ amount)."
            )
            raise ValueError(msg)
        found_sheet, found_header = discovered
        logger.info(
            "Detected portfolio in %s on sheet %r at header row %d",
            path.name,
            found_sheet,
            found_header,
        )
        df = _read_sheet_to_df(workbook, found_sheet, found_header)
        return _parse_portfolio_df(
            df,
            nav=nav,
            source=f"{path}::{found_sheet}",
            price_fetcher=price_fetcher,
        )
    finally:
        workbook.close()


def load_from_csv_string(csv_string: str, nav: float = 3_000_000_000.0) -> Portfolio:
    """Load a portfolio from a CSV string (for testing or direct input)."""
    import io

    df = pl.read_csv(io.StringIO(csv_string), infer_schema_length=1000)
    return _parse_portfolio_df(df, nav=nav, source="csv_string")


def _parse_portfolio_df(
    df: pl.DataFrame,
    nav: float = 3_000_000_000.0,
    source: str = "unknown",
    price_fetcher: PriceFetcher | None = None,
) -> Portfolio:
    """Parse a polars DataFrame into a Portfolio."""
    df, column_scales = _standardize_columns(df)

    # Validate required columns
    has_ticker = "ticker" in df.columns
    has_side = "side" in df.columns
    has_shares = "shares" in df.columns
    has_weight = "weight" in df.columns
    has_notional = "notional_usd" in df.columns
    has_rsi = "rsi" in df.columns
    has_beta = "beta" in df.columns

    if not has_ticker:
        msg = f"Missing required column 'ticker' in {source}. Found: {df.columns}"
        raise ValueError(msg)
    if not has_side and not has_shares and not has_notional:
        msg = (
            f"Missing required column 'side' in {source}. "
            f"Provide a 'side' column or use signed shares/notional. Found: {df.columns}"
        )
        raise ValueError(msg)
    if not has_shares and not has_weight and not has_notional:
        msg = (
            f"Need 'shares', 'weight', or a dollar-amount column "
            f"(e.g. '$ amount', 'notional', 'market_value') in {source}. Found: {df.columns}"
        )
        raise ValueError(msg)

    notional_scale = column_scales.get("notional_usd", 1.0)
    shares_scale = column_scales.get("shares", 1.0)
    entry_price_scale = column_scales.get("entry_price", 1.0)
    weight_scale = column_scales.get("weight", 1.0)

    # Always batch-fetch current prices when a fetcher is available. We need them
    # for two things: (1) converting $-notional → shares when no shares column is
    # supplied, and (2) populating Position.current_price so notional / gross /
    # net / market-value math actually has a price to multiply shares by. Without
    # this, current_price defaults to 0.0 and every downstream exposure metric
    # collapses to zero — even when shares + entry_price are both present.
    fetched_prices: dict[str, float] = {}
    if price_fetcher is not None:
        tickers_needing_prices = sorted(
            {
                _clean_ticker(str(row["ticker"]))
                for row in df.iter_rows(named=True)
                if row.get("ticker") is not None
                and not (isinstance(row.get("ticker"), str) and not row["ticker"].strip())
            }
        )
        if tickers_needing_prices:
            try:
                fetched_prices = price_fetcher(tickers_needing_prices) or {}
            except (ValueError, KeyError, RuntimeError, ConnectionError) as exc:
                logger.warning("Price fetch failed during ingest: %s", exc)

    positions: list[Position] = []

    for row in df.iter_rows(named=True):
        raw_ticker = row.get("ticker")
        if raw_ticker is None or (isinstance(raw_ticker, str) and not raw_ticker.strip()):
            # Footer rows like 'Total Portfolio Summary' have no ticker; skip silently
            continue
        ticker = _clean_ticker(str(raw_ticker))

        # Entry price (optional, used for $-to-shares conversion when notional supplied)
        entry_price = float(row.get("entry_price", 0) or 0) * entry_price_scale

        # Sizing precedence: explicit shares > $ notional > weight
        if has_shares and row.get("shares") is not None:
            shares = float(row["shares"]) * shares_scale
        elif has_notional and row.get("notional_usd") is not None:
            notional = float(row["notional_usd"]) * notional_scale
            price = entry_price if entry_price > 0 else fetched_prices.get(ticker, 0.0)
            if price <= 0:
                logger.warning(
                    "Dollar amount supplied for %s but no price (entry_price or fetched) "
                    "to convert to shares, skipping",
                    ticker,
                )
                continue
            shares = notional / price  # sign preserved for side inference
            # Preserve the resolved price so downstream exposure math (shares × entry_price)
            # reproduces the original notional. Without this, entry_price falls back to the
            # $1 placeholder and gross/net exposure collapses by a factor of price.
            if entry_price <= 0:
                entry_price = price
        elif has_weight and row.get("weight") is not None:
            weight = float(row["weight"]) * weight_scale
            shares = abs(weight) * nav / 100.0
        else:
            logger.warning("No shares, notional, or weight for %s, skipping", ticker)
            continue

        if shares == 0:
            logger.warning("Zero shares for %s, skipping", ticker)
            continue

        # Side: explicit column takes priority; otherwise infer from sign
        if has_side:
            side = _standardize_side(str(row["side"]))
        else:
            side = "SHORT" if shares < 0 else "LONG"
            shares = abs(shares)

        # Entry date (optional)
        entry_date_raw = row.get("entry_date")
        if entry_date_raw is not None and entry_date_raw != "":
            if isinstance(entry_date_raw, date):
                entry_date = entry_date_raw
            else:
                try:
                    entry_date = date.fromisoformat(str(entry_date_raw))
                except ValueError:
                    entry_date = date.today()
        else:
            entry_date = date.today()

        # Sector classification
        sector_override = str(row.get("sector", "") or "")
        subsector_override = str(row.get("subsector", "") or "")
        sector, subsector = classify_ticker(ticker)
        if sector_override:
            sector = sector_override
        if subsector_override:
            subsector = subsector_override

        # Asset type (optional — defaults to EQUITY)
        asset_type_raw = str(row.get("asset_type", "") or "").strip().upper()
        asset_type = asset_type_raw if asset_type_raw in ("EQUITY", "ETF", "OPTION") else "EQUITY"

        # Optional user-supplied RSI / beta — stored on Position, not consumed by metric modules
        rsi_val = row.get("rsi") if has_rsi else None
        beta_val = row.get("beta") if has_beta else None
        rsi_clean = float(rsi_val) if rsi_val is not None and rsi_val != "" else None
        beta_clean = float(beta_val) if beta_val is not None and beta_val != "" else None

        # Current price: prefer freshly fetched, then entry_price, then $1 placeholder
        current_price = fetched_prices.get(ticker, 0.0)
        if current_price <= 0:
            current_price = entry_price if entry_price > 0 else 1.0

        positions.append(
            Position(
                ticker=ticker,
                side=side,
                shares=shares,
                entry_price=entry_price if entry_price > 0 else current_price,
                current_price=current_price,
                entry_date=entry_date,
                sector=sector,
                subsector=subsector,
                asset_type=asset_type,
                rsi=rsi_clean,
                beta=beta_clean,
            )
        )

    if not positions:
        msg = f"No valid positions found in {source}"
        raise ValueError(msg)

    return Portfolio(
        name=Path(source).stem if source != "csv_string" else "Imported",
        positions=positions,
        nav=nav,
    )


def load_from_pdf(file_path: str | Path, nav: float = 3_000_000_000.0) -> Portfolio:
    """
    Best-effort portfolio extraction from PDF.

    Uses pdfplumber to find tables with ticker/position data.
    Falls back to text extraction if table extraction fails.
    """
    import pdfplumber

    path = Path(file_path)
    all_tables: list[list[list[str]]] = []

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            if tables:
                all_tables.extend(tables)

    if not all_tables:
        msg = f"No tables found in PDF: {path}. Please use Excel/CSV format instead."
        raise ValueError(msg)

    # Try to find a table that looks like a portfolio (has ticker-like data)
    for table in all_tables:
        if len(table) < 2:
            continue

        # Use first row as headers, rest as data
        headers = [str(h).strip() for h in table[0]]
        rows = table[1:]

        try:
            col_data = {
                headers[i]: [row[i] if i < len(row) else None for row in rows]
                for i in range(len(headers))
            }
            df = pl.DataFrame(col_data)
            return _parse_portfolio_df(df, nav=nav, source=str(path))
        except (ValueError, KeyError):
            continue

    msg = f"Could not find portfolio data in any PDF table from {path}"
    raise ValueError(msg)
